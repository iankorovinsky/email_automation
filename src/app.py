import time
from flask import Flask, render_template, request, redirect
from flask_mail import Mail, Message
from openpyxl import load_workbook, Workbook


app = Flask(__name__)
mail = Mail(app)

app.config['MAIL_SERVER'] = 'insert mail server here'
app.config['MAIL_PORT'] = <insert port number here>
app.config['MAIL_USERNAME'] = '<REPLACE_THIS_FROM_EMAIL_ADDRESS>'
app.config['MAIL_PASSWORD'] = '<REPLACE_THIS_EMAIL_PASSWORD>'
app.config['MAIL_USE_TLS'] = False
app.config['MAIL_USE_SSL'] = True
mail = Mail(app)
excelfile = 'directory.xlsx'
wb = load_workbook(excelfile)
sheet_obj = wb.active

@app.route("/", methods=["GET", "POST"])
def index():
    r = 2
    rows = sheet_obj.max_row
    transcript = ""
    if request.method == "POST":
        print("FORM DATA RECEIVED")
        for i in range(2, rows + 1):
            cName = sheet_obj.cell(row=r, column=1)
            pName = sheet_obj.cell(row=r, column=3)
            email = sheet_obj.cell(row=r, column=5)
            r = r + 1
            transcript = 'add message here'
            msg = Message('add subject line here', sender='###', recipients=[email.value], cc=['###'], bcc=['###'])
            msg.body = transcript
            with app.open_resource("attachment.pdf") as fp:
                msg.attach("attachment.pdf", "application/pdf", fp.read())
            mail.send(msg)
            print("Sent")
            time.sleep(600)

    return render_template('index.html', transcript=transcript)

if __name__ == "__main__":
    app.run(debug=True, threaded=True)

